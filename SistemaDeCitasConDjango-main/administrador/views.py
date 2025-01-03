from django.shortcuts import get_object_or_404, render, redirect
from .forms import (
    FormProfesional,
    FormLugares,
    FormHoras,
    PacienteForm,
    FormServicios,
    CargarHorarioArchivoForm,
)
from .models import (
    Horas,
    Lugares,
    Pacientes,
    Consultorio,
    Auditoria,
    Profesional,
    Servicio,
)
from usuario import models as modelsUsuario
from entrarSistema import forms as formEntrarSistema
from entrarSistema import models as modelsEntrarSistema
from django.contrib import messages
from django.http import JsonResponse
from django.db import IntegrityError
from django.utils.timezone import now
from django.core.paginator import Paginator
import openpyxl

# PDF
from django.http import HttpResponse
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch, cm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, Table, TableStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib import colors
from datetime import datetime

import xlrd
from django.shortcuts import render
from django.http import JsonResponse
from .forms import CargarHorarioArchivoForm


import openpyxl
from django.http import JsonResponse
from datetime import datetime, time


def cargar_horario(request):
    if request.method == "POST":
        form = CargarHorarioArchivoForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES.get("archivo")
            if archivo:
                if archivo.name.endswith(".xlsx"):
                    try:
                        # Usamos openpyxl para leer el archivo XLSX
                        workbook = openpyxl.load_workbook(archivo)
                        sheet = workbook.active  # Usamos la primera hoja del libro

                        print(f"El archivo XLSX tiene {sheet.max_row} filas.")
                        for fila_num, row in enumerate(
                            sheet.iter_rows(values_only=True), start=1
                        ):
                            print(f"Procesando fila {fila_num}: {row}")
                            try:
                                # Validar la longitud de la fila
                                if len(row) < 6:
                                    print(
                                        f"Fila {fila_num} tiene menos columnas de las esperadas: {len(row)} columnas."
                                    )
                                    continue

                                id_prof = int(
                                    row[0]
                                )  # Asegúrate de que el ID del profesional esté correctamente asignado como entero
                                hora_inicio = row[
                                    1
                                ]  # La hora de inicio ya está en formato time
                                hora_final = row[
                                    2
                                ]  # La hora de fin ya está en formato time
                                fecha_habilitada = row[
                                    3
                                ].date()  # Convertir fecha a formato adecuado
                                estado_hora = (
                                    True
                                    if row[4].strip().lower() == "activo"
                                    else False
                                )  # Convertir estado a booleano

                                # Obtener la instancia del profesional usando el ID
                                try:
                                    profesional = Profesional.objects.get(
                                        id_prof=id_prof
                                    )  # Obtener el profesional por su ID
                                except Profesional.DoesNotExist:
                                    print(
                                        f"El profesional con ID {id_prof} no existe en la base de datos."
                                    )
                                    continue  # Si el profesional no existe, pasa a la siguiente fila

                                # Reemplazar el registro si ya existe
                                Horas.objects.update_or_create(
                                    id_prof=profesional,
                                    inicio_hora=hora_inicio,
                                    final_hora=hora_final,
                                    fecha_habilitada=fecha_habilitada,
                                    defaults={
                                        "horas_estado": estado_hora  # Actualiza el estado
                                    },
                                )

                                print(
                                    f"Registro creado o actualizado para el profesional ID {id_prof}, Hora Inicio: {hora_inicio}, Hora Fin: {hora_final}, Fecha: {fecha_habilitada}, Estado: {estado_hora}"
                                )

                            except ValueError as ve:
                                print(f"Error de valor en la fila {row}: {ve}")
                            except Exception as e:
                                print(f"Error al procesar la fila {row}: {e}")

                        return JsonResponse(
                            {
                                "success": True,
                                "message": "Archivo procesado y datos guardados correctamente.",
                            }
                        )
                    except Exception as e:
                        print(f"Error al procesar el archivo: {str(e)}")
                        return JsonResponse(
                            {
                                "success": False,
                                "message": f"Error al procesar el archivo XLSX: {str(e)}",
                            }
                        )
                else:
                    return JsonResponse(
                        {
                            "success": False,
                            "message": "El archivo no es un archivo XLSX válido.",
                        }
                    )
            else:
                return JsonResponse(
                    {"success": False, "message": "No se ha recibido ningún archivo."}
                )
        else:
            return JsonResponse(
                {
                    "success": False,
                    "message": "Formulario no válido.",
                    "errors": form.errors,
                }
            )
    else:
        form = CargarHorarioArchivoForm()
    return render(request, "cargar_horario.html", {"form": form})


# Create your views here.


def inicioAdmin(request):

    # Si ya tiene sesión no le abre esta página
    user = request.user
    if not user.is_authenticated:
        # HttpResponse('<script>alert("funcionó");</script>')
        return redirect("iniciarSesion")

    tipoRol = TipoRol(request)

    if not tipoRol.es_administrador:
        return redirect("inicio")
    else:
        lista_citas = (
            Consultorio.objects.filter(id_cit__citas_estado=True)
            .exclude(id_cit__estado_cita="Rechazada")
            .distinct("id_cit")
            .order_by("-id_cit")
        )

        listaRoles = modelsEntrarSistema.UsuarioRoles.objects.filter(
            id_usu=request.user.id
        )

        return render(
            request,
            "inicio_admin.html",
            {"listaCitas": lista_citas, "listaRoles": listaRoles},
        )


def pdfCitas(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename= Lista-Citas-reporte.pdf"

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)

    # Cabecera
    c.setLineWidth(0.3)
    c.setFont("Helvetica", 22)
    c.drawString(30, 745, "Citas")
    c.setFont("Helvetica", 12)
    c.drawString(30, 725, "Reporte de citas realizadas")

    hoy_fecha = str(datetime.today().strftime("%Y-%m-%d"))
    hoy = "Fecha de Hoy: " + hoy_fecha

    c.setFont("Helvetica-Bold", 12)
    c.drawString(420, 750, hoy)
    c.line(415, 745, 570, 745)

    # Nombra de las listas
    styles = getSampleStyleSheet()
    styleBH = styles["Normal"]
    styleBH.alignment = TA_CENTER
    styleBH.fontSize = 10

    id = Paragraph("""#""", styleBH)
    representante = Paragraph("""REPRESENTANTE""", styleBH)
    paciente = Paragraph("""PACIENTE""", styleBH)
    lugar = Paragraph("""LUGAR""", styleBH)
    fecha = Paragraph("""FECHA""", styleBH)
    estado = Paragraph("""ESTADO""", styleBH)

    data = []
    data.append([id, representante, paciente, lugar, fecha, estado])

    # consultas = [cita.id_pac for cita in lista_select]

    lista_consulta = (
        Consultorio.objects.filter(id_cit__estado_cita="Realizada")
        .distinct("id_cit")
        .order_by("-id_cit")
    )

    # Configuracion del contenido de la tabla
    styleN = styles["Normal"]
    styleN.alignment = TA_LEFT
    styleN.fontSize = 11
    styleN.wordWrap = "LTR"

    high = 600
    for consulta in lista_consulta:

        if consulta.id_cit.id_pac:
            paciente = consulta.id_cit.id_pac.nombre_pac
        else:
            paciente = "No enlazado"

        id_cita = str(consulta.id_cit.id_cit)
        representante = str(consulta.id_cit.id_usu.nombre)

        pacientes = str(paciente)

        lugar = str(consulta.id_cit.id_lugar.ubicacion_lugar)
        fecha = str(consulta.id_cit.dia_cit)

        estado_cita = str(consulta.id_cit.estado_cita)

        this_estudiante = [
            Paragraph(id_cita, styleBH),
            Paragraph(representante, styleBH),
            Paragraph(pacientes, styleBH),
            Paragraph(lugar, styleBH),
            Paragraph(fecha, styleBH),
            Paragraph(estado_cita, styleBH),
        ]
        data.append(this_estudiante)
        high = high - 18

    # Contenido de la tabla
    width, height = A4
    table = Table(
        data,
        colWidths=[  # estilo de la tabla
            1.5 * cm,
            4.5 * cm,
            4.5 * cm,
            3.9 * cm,
            2.5 * cm,
            2.5 * cm,
        ],
    )
    table.setStyle(
        TableStyle(
            [
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
                ("BOX", (0, 0), (-1, -1), 0.25, colors.black),
            ]
        )
    )

    # tamaño del PDF
    table.wrapOn(c, width, height)
    table.drawOn(c, 30, high)
    c.showPage()  # guardar pagina

    # Guardar PDF
    c.save()
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


def lista_pacientes(request):
    # Manejar el formulario de creación de paciente
    if request.method == "POST":
        form = PacienteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "¡Paciente creado exitosamente!")
            return redirect("lista_pacientes")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error en {field}: {error}")
    else:
        form = PacienteForm()

    # Obtener todos los pacientes
    pacientes = Pacientes.objects.all()

    # Implementar la paginación
    paginator = Paginator(pacientes, 10)  # 10 pacientes por página
    page_number = request.GET.get("page")  # Obtener el número de página de la URL
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "pacientes.html",
        {
            "form": form,
            "page_obj": page_obj,  # Paginación de pacientes
        },
    )


def eliminar_paciente(request, id_pac):
    if request.method == "POST":
        # Obtener el objeto paciente
        paciente = get_object_or_404(Pacientes, id_pac=id_pac)

        # Verificar si el paciente tiene citas asociadas
        citas = modelsUsuario.Citas.objects.filter(id_pac=paciente)
        if citas.exists():
            # Si existen citas asociadas, devolver un mensaje de error
            return JsonResponse(
                {
                    "status": "error",
                    "message": "Este paciente tiene citas asociadas, no se puede eliminar.",
                }
            )

        # Si no tiene citas asociadas, eliminar el paciente
        paciente.delete()
        return JsonResponse(
            {"status": "success", "message": "Paciente eliminado exitosamente."}
        )

    return JsonResponse(
        {"status": "error", "message": "Error al procesar la solicitud."}
    )


# Vista para editar un paciente
def editar_paciente(request, id_pac):
    if request.method == "POST":
        # Obtener los datos enviados en la solicitud
        nombre_pac = request.POST.get("nombre_pac")
        nacimiento_pac = request.POST.get("nacimiento_pac")
        genero_pac = request.POST.get("genero_pac")
        estado_pac = request.POST.get("pacientes_estado")

        # Convertir la fecha en formato deseado
        nacimiento_pac = datetime.strptime(nacimiento_pac, "%Y-%m-%d").date()

        # Validar que la fecha de nacimiento no sea mayor a hoy
        if nacimiento_pac > datetime.today().date():
            return JsonResponse(
                {
                    "status": "error",
                    "message": "La fecha de nacimiento no puede ser mayor a hoy.",
                }
            )

        # Obtener el paciente a editar usando id_pac como clave primaria
        paciente = Pacientes.objects.get(id_pac=id_pac)  # Cambié de id a id_pac

        # Actualizar los campos con los nuevos valores
        paciente.nombre_pac = nombre_pac
        paciente.nacimiento_pac = nacimiento_pac
        paciente.genero_pac = genero_pac
        paciente.estado_pac = estado_pac

        # Guardar los cambios en la base de datos
        paciente.save()

        return JsonResponse({"status": "success"})
    else:
        paciente = Pacientes.objects.get(id_pac=id_pac)  # Cambié de id a id_pac
        formatted_nacimiento = paciente.nacimiento_pac.strftime(
            "%Y-%m-%d"
        )  # Formato adecuado para el input de tipo 'date'

        return render(
            request,
            "editar_paciente.html",
            {
                "paciente": paciente,
                "formatted_nacimiento": formatted_nacimiento,
            },
        )


def is_ajax(request):
    return request.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest"


from datetime import datetime

def consultorio_cita(request):
    user = request.user
    if not user.is_authenticated:
        return redirect("iniciarSesion")

    tipoRol = TipoRol(request)

    if not tipoRol.es_administrador:
        return redirect("inicio")
    else:
        if request.method == "GET":
            # Obtener la cita seleccionada
            lista_select = modelsUsuario.Citas.objects.filter(id_cit=request.GET["id_cit"])

            consultas = [cita.id_pac for cita in lista_select]
            if None in consultas:
                lista_consulta = ""
            else:
                # Filtrar las consultas previas por paciente y estado "Realizada"
                lista_consulta = Consultorio.objects.filter(
                    id_cit__id_pac=consultas[0], consultorio_estado=True
                )

            return render(
                request,
                "consultorio_citas.html",
                {"citaSelect": lista_select, "listaPrevia": lista_consulta},
            )

        else:
            # Obtener la cita de la base de datos
            instance_Citas = modelsUsuario.Citas.objects.get(id_cit=request.POST["id_cit"])

            # Validar y convertir el campo nacimiento_con
            nacimiento_hidden = request.POST.get("nacimiento_con")
            nacimiento_convertido = None
            if nacimiento_hidden and nacimiento_hidden != "nada":
                try:
                    nacimiento_convertido = datetime.strptime(nacimiento_hidden, "%d de %B de %Y").date()
                except ValueError:
                    return render(
                        request,
                        "consultorio_citas.html",
                        {
                            "error": "El formato de la fecha de nacimiento no es válido. Debe ser YYYY-MM-DD.",
                            "citaSelect": modelsUsuario.Citas.objects.filter(id_cit=request.POST["id_cit"]),
                        },
                    )

            # Crear la instancia de Consultorio
            idCon = Consultorio(
                id_cit=instance_Citas,
                nota_con=request.POST["nota_con"],
                nacimiento_con=nacimiento_convertido,
            )
            idCon.save()

            # Actualizar el estado de la cita a "Realizada"
            idCit = modelsUsuario.Citas.objects.get(id_cit=request.POST["id_cit"])
            idCit.estado_cita = "Realizada"
            idCit.save()

            # Registrar la auditoría
            Audi = Auditoria(
                descripcion_aut=f"Se creó una 'consulta' en la tabla *Consultorio*, con la nota de la consulta ({request.POST['nota_con']}) con el id {idCon.pk}, creado por el usuario: {request.user.id},"
            )
            Audi.save()

            return redirect("consultorio")


def consultorio(request):
    # Verificar si el usuario está autenticado
    user = request.user
    if not user.is_authenticated:
        return redirect("iniciarSesion")

    fecha_hoy = datetime.today().strftime("%Y-%m-%d")
    tipoRol = TipoRol(request)

    if not tipoRol.es_administrador:
        return redirect("inicio")
    else:
        if request.method == "GET":
            # Obtener las citas de hoy con estado "Aceptada"
            lista_citas = modelsUsuario.Citas.objects.filter(
                citas_estado=True, estado_cita="Aceptada", dia_cit=datetime.today()
            ).order_by("-id_cit")

            # Obtener las citas de todos los días con estado "Aceptada"
            lista_consultorio = modelsUsuario.Citas.objects.filter(
                citas_estado=True, estado_cita="Aceptada"
            )

            return render(
                request,
                "consultorio.html",
                {
                    "listaCitas": lista_citas,
                    "citaSelect": modelsUsuario.Citas.objects.filter(id_cit=9),  # Es posible que quieras usar otro criterio aquí
                    "listaConsultorio": lista_consultorio,
                    "fecha": fecha_hoy,
                },
            )
        else:
            # Obtener las citas y consultorios
            lista_citas = modelsUsuario.Citas.objects.filter(citas_estado=True, estado_cita="Aceptada")
            lista_consultorio = modelsUsuario.Citas.objects.filter(citas_estado=True, estado_cita="Aceptada")

            instance_Citas = modelsUsuario.Citas.objects.get(id_cit=request.POST["id_cit"])

            # Crear la instancia de Consultorio
            idCon = Consultorio(
                id_cit=instance_Citas,
                nota_con=request.POST["nota_con"],
                nacimiento_con=request.POST["nacimiento_con"],  # Usar el campo nacimiento_con
            )
            idCon.save()

            # Actualizar el estado de la cita a "Realizada"
            idCit = modelsUsuario.Citas(id_cit=request.POST["id_cit"])
            idCit.estado_cita = "Realizada"
            idCit.save()

            # Registrar la auditoría
            Audi = Auditoria(
                descripcion_aut=f"Se creó una 'consulta' en la tabla *Consultorio*, con la nota de la consulta ({request.POST['nota_con']}) con el id {idCon.pk}, creado por el usuario: {request.user.id},"
            )
            Audi.save()

            return render(
                request,
                "consultorio.html",
                {
                    "listaCitas": lista_citas,
                    "citaSelect": modelsUsuario.Citas.objects.filter(id_cit=9),
                    "listaConsultorio": lista_consultorio,
                    "fecha": fecha_hoy,
                },
            )



def citas_admin(request):
    # Verificar si el usuario está autenticado
    user = request.user
    if not user.is_authenticated:
        return redirect("iniciarSesion")

    tipoRol = TipoRol(request)

    if not tipoRol.es_administrador:
        return redirect("inicio")
    else:
        if request.method == "GET":
            lista_citas_aceptadas = modelsUsuario.Citas.objects.filter(
                estado_cita="Aceptada", citas_estado=True
            )
            lista_citas_rechazadas = modelsUsuario.Citas.objects.filter(
                estado_cita="Rechazada", citas_estado=True
            )
            lista_citas_pendientes = modelsUsuario.Citas.objects.filter(
                estado_cita="Sin confirmar", citas_estado=True
            )
            lista_citas_realizadas = modelsUsuario.Citas.objects.filter(
                estado_cita="Realizada", citas_estado=True
            )

            if request.is_ajax():  # Si es una solicitud AJAX, devolver las citas actualizadas
                citas_data = {
                    'aceptadas': list(lista_citas_aceptadas.values('id_cit', 'id_usu__nombre', 'id_pac__nombre_pac', 'dia_cit', 'id_hora__inicio_hora', 'id_hora__final_hora', 'id_lugar__nombre_lugar')),
                    'rechazadas': list(lista_citas_rechazadas.values('id_cit', 'id_usu__nombre', 'id_pac__nombre_pac', 'dia_cit', 'id_hora__inicio_hora', 'id_hora__final_hora', 'id_lugar__nombre_lugar')),
                    'pendientes': list(lista_citas_pendientes.values('id_cit', 'id_usu__nombre', 'id_pac__nombre_pac', 'dia_cit', 'id_hora__inicio_hora', 'id_hora__final_hora', 'id_lugar__nombre_lugar')),
                    'realizadas': list(lista_citas_realizadas.values('id_cit', 'id_usu__nombre', 'id_pac__nombre_pac', 'dia_cit', 'id_hora__inicio_hora', 'id_hora__final_hora', 'id_lugar__nombre_lugar')),
                }
                return JsonResponse(citas_data)

            # Si no es una solicitud AJAX, simplemente renderizar la vista
            return render(
                request,
                "citas_admin.html",
                {
                    "listaCitasAceptadas": lista_citas_aceptadas,
                    "listaCitasRechazadas": lista_citas_rechazadas,
                    "listaCitasPendientes": lista_citas_pendientes,
                    "listaCitasRealizadas": lista_citas_realizadas,
                },
            )
        else:
            # Aquí maneja los cambios de estado
            if "btnRevisarAceptar" in request.POST:
                idCita = modelsUsuario.Citas.objects.get(id_cit=request.POST["id_cit"])
                idCita.estado_cita = "Aceptada"
                idCita.save()
                # Registrar en la auditoría
                Audi = Auditoria(
                    descripcion_aut=f"Se aceptó una 'cita' con id {request.POST['id_cit']}, aceptado por el usuario: {request.user.id}"
                )
                Audi.save()

            elif "btnRevisarRechazar" in request.POST:
                idCita = modelsUsuario.Citas.objects.get(id_cit=request.POST["id_cit"])
                idCita.estado_cita = "Rechazada"
                idCita.save()
                # Registrar en la auditoría
                Audi = Auditoria(
                    descripcion_aut=f"Se rechazó una 'cita' con id {request.POST['id_cit']}, rechazado por el usuario: {request.user.id}"
                )
                Audi.save()

            elif "btnAceptadaRechazar" in request.POST:
                idCita = modelsUsuario.Citas.objects.get(id_cit=request.POST["id_cit"])
                idCita.estado_cita = "Rechazada"
                idCita.save()
                # Registrar en la auditoría
                Audi = Auditoria(
                    descripcion_aut=f"Se rechazó una cita aceptada con id {request.POST['id_cit']}, rechazado por el usuario: {request.user.id}"
                )
                Audi.save()

            elif "btnRechazarAceptar" in request.POST:
                idCita = modelsUsuario.Citas.objects.get(id_cit=request.POST["id_cit"])
                idCita.estado_cita = "Aceptada"
                idCita.save()
                # Registrar en la auditoría
                Audi = Auditoria(
                    descripcion_aut=f"Se aceptó una cita rechazada con id {request.POST['id_cit']}, aceptado por el usuario: {request.user.id}"
                )
                Audi.save()

            elif "btnRechazarEliminar" in request.POST:
                idCita = modelsUsuario.Citas.objects.get(id_cit=request.POST["id_cit"])
                idCita.citas_estado = False
                idCita.save()
                # Registrar en la auditoría
                Audi = Auditoria(
                    descripcion_aut=f"Se eliminó una cita con id {request.POST['id_cit']}, eliminado por el usuario: {request.user.id}"
                )
                Audi.save()

            # Después de cualquier cambio de estado, devolver las citas actualizadas por AJAX
            lista_citas_aceptadas = modelsUsuario.Citas.objects.filter(
                estado_cita="Aceptada", citas_estado=True
            )
            lista_citas_rechazadas = modelsUsuario.Citas.objects.filter(
                estado_cita="Rechazada", citas_estado=True
            )
            lista_citas_pendientes = modelsUsuario.Citas.objects.filter(
                estado_cita="Sin confirmar", citas_estado=True
            )
            lista_citas_realizadas = modelsUsuario.Citas.objects.filter(
                estado_cita="Realizada", citas_estado=True
            )

            if request.is_ajax():
                citas_data = {
                    'aceptadas': list(lista_citas_aceptadas.values('id_cit', 'id_usu__nombre', 'id_pac__nombre_pac', 'dia_cit', 'id_hora__inicio_hora', 'id_hora__final_hora', 'id_lugar__nombre_lugar')),
                    'rechazadas': list(lista_citas_rechazadas.values('id_cit', 'id_usu__nombre', 'id_pac__nombre_pac', 'dia_cit', 'id_hora__inicio_hora', 'id_hora__final_hora', 'id_lugar__nombre_lugar')),
                    'pendientes': list(lista_citas_pendientes.values('id_cit', 'id_usu__nombre', 'id_pac__nombre_pac', 'dia_cit', 'id_hora__inicio_hora', 'id_hora__final_hora', 'id_lugar__nombre_lugar')),
                    'realizadas': list(lista_citas_realizadas.values('id_cit', 'id_usu__nombre', 'id_pac__nombre_pac', 'dia_cit', 'id_hora__inicio_hora', 'id_hora__final_hora', 'id_lugar__nombre_lugar')),
                }
                return JsonResponse(citas_data)

        # Si no es AJAX, solo renderizar la vista
        return render(
            request,
            "citas_admin.html",
            {
                "listaCitasAceptadas": lista_citas_aceptadas,
                "listaCitasRechazadas": lista_citas_rechazadas,
                "listaCitasPendientes": lista_citas_pendientes,
                "listaCitasRealizadas": lista_citas_realizadas,
            },
        )


def cfg_admin(request):
    # Si ya tiene sesión no le abre esta página
    user = request.user
    if not user.is_authenticated:
        # HttpResponse('<script>alert("funcionó");</script>')
        return redirect("iniciarSesion")

    tipoRol = TipoRol(request)

    if not tipoRol.es_usuario:
        return redirect("inicio")
    else:
        form = formEntrarSistema.FormRegistrar()
        if request.method == "GET":
            form = formEntrarSistema.FormRegistrar()
            return render(request, "configuracion_admin.html", {"form": form})
        else:

            idUser = modelsEntrarSistema.CrearCuenta.objects.get(id=request.user.id)

            if "btnUsuario" in request.POST:

                if modelsEntrarSistema.CrearCuenta.objects.filter(
                    username=request.POST["username"]
                ).exists():
                    HttpResponse("<script>alert('Ya existe ese usuario')</script>")

                    return render(request, "configuracion_admin.html", {"form": form})
                else:
                    idUser.username = request.POST["username"]
                    idUser.save()

                    # Aqui ponemos el codigo del trigger -------

                    Audi = Auditoria(
                        descripcion_aut=f"Se modifico una 'cuenta' en la tabla *CrearCuenta*, se cambio el usuario por {request.POST['username']}."
                    )
                    Audi.save()

                    # fin de trigger ------

                    return render(request, "configuracion_admin.html", {"form": form})

            elif "btnCedula" in request.POST:
                if modelsEntrarSistema.CrearCuenta.objects.filter(
                    cedula=request.POST["cedula"]
                ).exists():
                    HttpResponse(
                        "<script>alert('Ya se ha registrado esta cédula')</script>"
                    )

                    return render(request, "configuracion_admin.html", {"form": form})
                else:
                    idUser.cedula = request.POST["cedula"]
                    idUser.save()
                    # Aqui ponemos el codigo del trigger -------

                    Audi = Auditoria(
                        descripcion_aut=f"Se modifico una 'cuenta' en la tabla *CrearCuenta*, se cambio la cedula, del usuario {request.user.id} ."
                    )
                    Audi.save()

                    # fin de trigger ------

                    return render(request, "configuracion_admin.html", {"form": form})

            elif "btnTelefono" in request.POST:

                if not request.POST["numero"]:
                    HttpResponse(
                        "<script>alert('Tiene que escribir un número telefónico')</script>"
                    )

                    return render(request, "configuracion_admin.html", {"form": form})
                else:
                    idUser.numero = request.POST["numero"]
                    idUser.save()

                    # Aqui ponemos el codigo del trigger -------

                    Audi = Auditoria(
                        descripcion_aut=f"Se modifico una 'cuenta' en la tabla *CrearCuenta*, se cambio el numero de teléfono, del usuario {request.user.id} ."
                    )
                    Audi.save()

                    # fin de trigger ------

                    return render(request, "configuracion_admin.html", {"form": form})

            elif "btnCorreo" in request.POST:

                if modelsEntrarSistema.CrearCuenta.objects.filter(
                    correo=request.POST["correo"]
                ).exists():
                    HttpResponse("<script>alert('Ya existe este correo')</script>")

                    return render(request, "configuracion_admin.html", {"form": form})
                else:
                    idUser.correo = request.POST["correo"]
                    idUser.save()
                    print(idUser.save())

                    # Aqui ponemos el codigo del trigger -------

                    Audi = Auditoria(
                        descripcion_aut=f"Se modifico una 'cuenta' en la tabla *CrearCuenta*, se cambio el correo electrónico por {request.user.id}."
                    )
                    Audi.save()

                    # fin de trigger ------

                    return render(request, "configuracion_admin.html", {"form": form})

            elif "btnContrasenna" in request.POST:

                if not request.POST["password1"] == request.POST["password2"]:
                    HttpResponse(
                        "<script>alert('Las contraseñas deben ser iguales')</script>"
                    )

                    return render(request, "configuracion_admin.html", {"form": form})
                else:
                    idUser.set_password(request.POST["password1"])
                    idUser.save()
                    # Aqui ponemos el codigo del trigger -------

                    Audi = Auditoria(
                        descripcion_aut=f"Se modifico una 'cuenta' en la tabla *CrearCuenta*, se cambio la contraseña por {request.user.id}."
                    )
                    Audi.save()

                    # fin de trigger ------

                    return render(request, "configuracion_admin.html", {"form": form})

            elif "btnAll" in request.POST:
                if (
                    modelsEntrarSistema.CrearCuenta.objects.filter(
                        correo=request.POST["correo"]
                    ).exists()
                    or modelsEntrarSistema.CrearCuenta.objects.filter(
                        username=request.POST["username"]
                    ).exists()
                    or modelsEntrarSistema.CrearCuenta.objects.filter(
                        cedula=request.POST["cedula"]
                    ).exists()
                ):
                    HttpResponse(
                        "<script>alert('Los datos suministrados ya existen en el sistema, elija otros por favor.')</script>"
                    )
                    return render(request, "configuracion_admin.html", {"form": form})
                else:
                    if request.POST["password1"] == request.POST["password2"]:

                        idUser.username = request.POST["username"]
                        idUser.cedula = request.POST["cedula"]
                        idUser.numero = request.POST["numero"]
                        idUser.correo = request.POST["correo"]
                        idUser.set_password(request.POST["password1"])
                        idUser.save()

                        # Aqui ponemos el codigo del trigger -------

                        Audi = Auditoria(
                            descripcion_aut=f"Se modifico una 'cuenta' en la tabla *CrearCuenta*, se cambio todos los datos por {request.user.id}."
                        )
                        Audi.save()

                        # fin de trigger ------

                        return render(
                            request, "configuracion_admin.html", {"form": form}
                        )

                    else:
                        HttpResponse(
                            "<script>alert('Las contraseñas deben ser iguales')</script>"
                        )
                        return render(
                            request, "configuracion_admin.html", {"form": form}
                        )

        return render(request, "configuracion_admin.html", {"form": form})


# CRUD DE CONFIGURACION ADMIN


def TipoRol(request):
    instanceHoras = formEntrarSistema.UsuarioRoles.objects.get(id_usu=request.user)

    return instanceHoras


# AQUI TENEMOS LAS VISTAS DE PROFESIONAL-------


def lista_profesionales(request):
    if request.method == "POST":
        if "submitProfesional" in request.POST:
            formProfesional = FormProfesional(request.POST)
            if formProfesional.is_valid():
                try:
                    # Crear el profesional
                    nuevo_profesional = formProfesional.save(commit=False)
                    lugares_seleccionados = formProfesional.cleaned_data.get("lugares")
                    
                    # Verificar si alguno de los lugares ya está asociado a otro profesional
                    lugares_ocupados = Profesional.objects.filter(
                        lugares__in=lugares_seleccionados
                    ).distinct()
                    
                    if lugares_ocupados.exists():
                        # Obtener los nombres de los lugares ocupados
                        lugar_nombres = ", ".join(
                            lugar for lugar in lugares_ocupados.values_list(
                                "lugares__nombre_lugar", flat=True
                            )
                        )
                        formProfesional.add_error(
                            "lugares",
                            f"Los lugares '{lugar_nombres}' ya están asociados a otros profesionales."
                        )
                        return render(
                            request,
                            "profesionales.html",
                            {
                                "formProfesional": formProfesional,
                                "profesionales": Profesional.objects.all(),
                                "lugares": Lugares.objects.all(),
                            },
                        )
                    
                    # Guardar el profesional y asignar los lugares seleccionados
                    nuevo_profesional.save()  # Guarda primero el profesional
                    if lugares_seleccionados:
                        nuevo_profesional.lugares.set(lugares_seleccionados)  # Relacionar lugares
                    return redirect("lista_profesionales")
                
                except IntegrityError:
                    formProfesional.add_error(
                        "num_doc_prof", "El número de documento ya existe."
                    )
            else:
                return render(
                    request,
                    "profesionales.html",
                    {
                        "formProfesional": formProfesional,
                        "profesionales": Profesional.objects.all(),
                        "lugares": Lugares.objects.all(),
                    },
                )

    # GET Request
    formProfesional = FormProfesional()
    profesionales = Profesional.objects.all()
    lugares = Lugares.objects.all()
    lugares_ids = lugares.values_list(
        "id_lugar", flat=True
    )  # Obtener los ids de los lugares
    return render(
        request,
        "profesionales.html",
        {
            "formProfesional": formProfesional,
            "profesionales": profesionales,
            "lugares": lugares,
            "lugares_ids": lugares_ids,  # Pasa los IDs de los lugares al contexto
        },
    )




def editar_profesional(request, id_prof):
    if request.method == "POST":
        # Obtener el profesional a editar
        profesional = get_object_or_404(Profesional, id_prof=id_prof)

        # Actualizar los campos básicos
        profesional.nombre_prof = request.POST.get("nombre_prof", profesional.nombre_prof)
        profesional.especialidad_prof = request.POST.get(
            "especialidad_prof", profesional.especialidad_prof
        )
        profesional.email_prof = request.POST.get("email_prof", profesional.email_prof)
        profesional.telefono_prof = request.POST.get("telefono_prof", profesional.telefono_prof)
        profesional.estado_prof = request.POST.get("estado_prof") == "True"  # Convertir a booleano

        # Obtener los lugares seleccionados (IDs enviados desde el formulario)
        lugares_ids = request.POST.getlist("lugares[]")
        print("Lugares seleccionados:", lugares_ids)  # Verifica si los lugares están llegando correctamente

        if lugares_ids:
            # Validar si algún lugar ya está asociado a otro profesional
            lugares = Lugares.objects.filter(id_lugar__in=lugares_ids)
            for lugar in lugares:
                # Verificar si algún otro profesional ya está asociado al lugar
                otros_profesionales = lugar.profesionales.exclude(id_prof=id_prof)
                if otros_profesionales.exists():
                    return JsonResponse(
                        {
                            "success": False,
                            "message": f"El lugar '{lugar.nombre_lugar}' ya está asociado a otro profesional.",
                        }
                    )

            # Actualizar la relación ManyToMany si pasa la validación
            profesional.lugares.set(lugares)
        else:
            profesional.lugares.clear()  # Si no hay lugares seleccionados, limpia la relación

        # Guardar los cambios
        profesional.save()

        return JsonResponse({"success": True, "message": "Profesional editado correctamente."})
    else:
        return JsonResponse({"status": "error", "message": "Método no permitido."})


def eliminar_profesional(request, id_prof):
    if request.method == "POST":
        # Obtener el objeto profesional
        profesional = get_object_or_404(Profesional, id_prof=id_prof)

        # Verificar si el profesional tiene citas asociadas
        citas = modelsUsuario.Citas.objects.filter(id_prof=profesional)
        if citas.exists():
            # Si existen citas asociadas, devolver un mensaje de error
            return JsonResponse(
                {
                    "status": "error",
                    "message": "Este profesional tiene citas asociadas, no se puede eliminar."
                }
            )

        # Verificar si el profesional tiene horarios asociados
        horas = Horas.objects.filter(id_prof=profesional)
        if horas.exists():
            # Si existen horas asociadas, devolver un mensaje de error
            return JsonResponse(
                {
                    "status": "error",
                    "message": "Este profesional tiene horarios asociados, no se puede eliminar."
                }
            )

        try:
            # Si no tiene citas ni horas asociadas, proceder con la eliminación
            profesional.delete()
            return JsonResponse(
                {"status": "success", "message": "Profesional eliminado exitosamente."}
            )
        except IntegrityError as e:
            # Manejar cualquier error de integridad al eliminar
            return JsonResponse(
                {"status": "error", "message": f"Error de integridad: {str(e)}"}
            )
        except Exception as e:
            # Manejar otros errores inesperados
            return JsonResponse(
                {"status": "error", "message": f"Hubo un error al eliminar al profesional: {str(e)}"}
            )

    return JsonResponse(
        {"status": "error", "message": "Error al procesar la solicitud."}
    )

# AQUI TENEMOS LAS VISTAS DE LUGAR -------


def lista_lugar(request):
    if request.method == "POST":
        if "submitLugares" in request.POST:
            formLugares = FormLugares(request.POST)
            if formLugares.is_valid():
                formLugares.save()
                return redirect("lista_lugar")
    else:
        formLugares = FormLugares()
        lugares = Lugares.objects.all()

    context = {
        "formLugares": formLugares,
        "lugares": lugares,
    }
    return render(request, "lugares.html", context)


def editar_lugar(request, id_lugar):
    if request.method == "POST":
        lugar = Lugares.objects.get(id_lugar=id_lugar)
        lugar.nombre_lugar = request.POST["nombre_lugar"]
        lugar.ubicacion_lugar = request.POST["ubicacion_lugar"]
        lugar.save()
        return JsonResponse({"status": "success"})
    return JsonResponse({"status": "error"})


def eliminar_lugar(request, id_lugar):
    if request.method == "POST":
        try:
            # Obtener el lugar por su ID
            lugar = Lugares.objects.get(id_lugar=id_lugar)

            # Verificar si el lugar está asociado a alguna cita
            if modelsUsuario.Citas.objects.filter(id_lugar=lugar).exists():
                return JsonResponse(
                    {
                        "status": "error",
                        "message": "No se puede eliminar el lugar, está asociado a una o más citas.",
                    }
                )

            # Verificar si el lugar está asociado a algún profesional
            if Profesional.objects.filter(lugares=lugar).exists():
                return JsonResponse(
                    {
                        "status": "error",
                        "message": "No se puede eliminar el lugar, está asociado a uno o más profesionales.",
                    }
                )

            # Si no está asociado a ningún profesional ni cita, proceder con la eliminación
            lugar.delete()
            return JsonResponse({"status": "success", "message": "Lugar eliminado exitosamente."})

        except Lugares.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Lugar no encontrado."})




# AQUI TENEMOS LAS VISTAS DE HORARIOS -------


def lista_horario(request):
    formHoras = FormHoras()  # Inicializar el formulario para horarios
    query = Horas.objects.all()  # Inicializar la consulta de horarios

    if request.method == "POST":
        if "archivoExcel" in request.FILES:  # Verificar si se subió un archivo
            archivo = request.FILES["archivoExcel"]
            # Procesar el archivo aquí
            # Por ejemplo, podrías validar el archivo
            try:
                # Procesamiento exitoso del archivo
                messages.success(request, "Archivo cargado correctamente.")
            except Exception as e:
                messages.error(request, f"Error al procesar el archivo: {str(e)}")
            return redirect("lista_horario")
        elif "submitHoras" in request.POST:  # Verificar si es el formulario de horarios
            formHoras = FormHoras(request.POST)
            if formHoras.is_valid():
                formHoras.save()
                messages.success(request, "Horario creado exitosamente.")
                return redirect("lista_horario")
            else:
                messages.error(request, "Corrige los errores en el formulario.")

        elif "submitHorasFiltro" in request.POST:
            # Filtrar por nombre profesional
            if "num_doc_prof" in request.POST and request.POST["num_doc_prof"]:
                print(f"Filtrando por num_doc_prof: {request.POST['num_doc_prof']}")
                query = query.filter(
                    id_prof__num_doc_prof__icontains=request.POST["num_doc_prof"]
                )

            # Filtrar por fecha habilitada
            if "fecha_habilitada" in request.POST and request.POST["fecha_habilitada"]:
                print(
                    f"Filtrando por fecha_habilitada: {request.POST['fecha_habilitada']}"
                )
                query = query.filter(fecha_habilitada=request.POST["fecha_habilitada"])

            # Filtrar por estado de la hora
            if "estado_hora" in request.POST and request.POST["estado_hora"]:
                estado_hora = request.POST["estado_hora"]
                # Convertir 'Activo' -> True, 'Inactivo' -> False
                if estado_hora == "Activo":
                    query = query.filter(horas_estado=True)
                elif estado_hora == "Inactivo":
                    query = query.filter(horas_estado=False)

            # Verifica qué contiene la query después de los filtros
            print(f"Consulta final: {query.query}")

    # Implementar la paginación
    paginator = Paginator(query, 10)  # Paginación sobre la query filtrada
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Contexto para la plantilla
    context = {
        "formHoras": formHoras,
        "page_obj": page_obj,
    }
    return render(request, "horarios.html", context)


def editar_horario(request, id_hora):
    if request.method == "POST":
        print(request.POST)  # Depuración: imprimir los datos recibidos

        try:
            horario = Horas.objects.get(id_hora=id_hora)
        except Horas.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Horario no encontrado"})

        # Actualizar horas de inicio y fin
        horario.inicio_hora = request.POST.get("inicio_hora", horario.inicio_hora)
        horario.final_hora = request.POST.get("final_hora", horario.final_hora)

        # Obtener y convertir el estado
        estado = request.POST.get("horas_estado", None)  # None si no se envía el estado
        print(f"Estado recibido: {estado}")  # Depuración

        if estado == "A":  # Activo
            horario.horas_estado = True
        elif estado == "I":  # Inactivo
            horario.horas_estado = False
        else:
            return JsonResponse(
                {"status": "error", "message": "Estado inválido"}, status=400
            )

        print(f"Estado final (booleano): {horario.horas_estado}")  # Depuración final

        # Validar y actualizar fecha habilitada
        try:
            fecha_habilitada = request.POST.get("fecha_habilitada")
            horario.fecha_habilitada = (
                datetime.strptime(fecha_habilitada, "%Y-%m-%d").date()
                if fecha_habilitada
                else horario.fecha_habilitada
            )
        except ValueError:
            return JsonResponse(
                {"status": "error", "message": "Fecha habilitada inválida"}
            )

        # Validar y asignar el profesional
        try:
            profesional = Profesional.objects.get(
                num_doc_prof=request.POST.get("num_doc_prof")
            )
            horario.id_prof = profesional
        except Profesional.DoesNotExist:
            return JsonResponse(
                {"status": "error", "message": "Profesional no encontrado"}
            )

        # Guardar los cambios
        horario.save()
        return JsonResponse(
            {"status": "success", "message": "Horario actualizado exitosamente."}
        )

    return JsonResponse({"status": "error", "message": "Método no permitido"})


def eliminar_horario(request, id_hora):
    if request.method == "POST":
        try:
            # Obtener el objeto horario
            horario = Horas.objects.get(id_hora=id_hora)

            # Comprobar si hay citas asociadas con este horario
            citas = modelsUsuario.Citas.objects.filter(id_hora=horario)
            if citas.exists():
                # Si existen citas asociadas, devolver un mensaje de error
                return JsonResponse(
                    {
                        "status": "error",
                        "message": "Este horario está asociado a citas, no se puede eliminar.",
                    }
                )

            # Si no hay citas asociadas, eliminar el horario
            horario.delete()
            return JsonResponse(
                {"status": "success", "message": "Horario eliminado exitosamente."}
            )

        except Horas.DoesNotExist:
            # Si no se encuentra el horario, devolver mensaje de error
            return JsonResponse({"status": "error", "message": "Horario no encontrado"})


# AQUI TENEMOS LAS VISTAS DE LA CREACION DE LOS SEVICIOS -------


def lista_servicios(request):
    if request.method == "POST":
        formServicios = FormServicios(request.POST)
        if formServicios.is_valid():
            # Verificar si ya existe un servicio con el mismo nombre
            nombre_servicio = formServicios.cleaned_data.get("nombre_servicio")
            if Servicio.objects.filter(nombre_servicio=nombre_servicio).exists():
                messages.error(
                    request,
                    f"El servicio con el nombre '{nombre_servicio}' ya existe. Por favor, elige otro nombre.",
                )
            else:
                # Crear el servicio sin guardar aún en la base de datos
                servicio = formServicios.save(commit=False)
                # Obtener los profesionales seleccionados
                profesionales = formServicios.cleaned_data.get("profesionales")
                # Guardar el servicio
                servicio.save()
                # Asociar los profesionales seleccionados al servicio
                servicio.profesionales.set(profesionales)
                # Confirmar éxito
                messages.success(request, "Servicio creado exitosamente.")
                return redirect("lista_servicios")
        else:
            # Si el formulario no es válido, mostrar errores
            messages.error(request, "Por favor corrija los errores en el formulario.")
    else:
        # Si es GET, instanciar un formulario vacío
        formServicios = FormServicios()

    # Obtener todos los servicios y profesionales para mostrarlos en la plantilla
    servicios = Servicio.objects.prefetch_related("profesionales").all()
    profesionales = Profesional.objects.all()

    return render(
        request,
        "servicios.html",
        {
            "formServicios": formServicios,
            "servicios": servicios,
            "profesionales": profesionales,
        },
    )


def eliminar_servicios(request, id_servicio):
    if request.method == "POST":
        # Obtener el objeto servicio
        servicio = get_object_or_404(Servicio, id_servicio=id_servicio)

        # Verificar si el servicio tiene citas asociadas
        citas = modelsUsuario.Citas.objects.filter(id_servicio=servicio)
        if citas.exists():
            # Si existen citas asociadas, devolver un mensaje de error
            return JsonResponse(
                {
                    "status": "error",
                    "message": "Este servicio tiene citas asociadas, no se puede eliminar.",
                }
            )

        # Si no tiene citas asociadas, eliminar el servicio
        servicio.delete()
        return JsonResponse(
            {"status": "success", "message": "Servicio eliminado exitosamente."}
        )

    return JsonResponse(
        {"status": "error", "message": "Error al procesar la solicitud."}
    )


def editar_servicios(request, id_servicio):
    if request.method == "POST":
        print("Datos POST recibidos:")
        print(request.POST)

        # Obtener el servicio a editar
        servicio = get_object_or_404(Servicio, id_servicio=id_servicio)

        # Actualizar los campos básicos del servicio
        servicio.nombre_servicio = request.POST.get("nombre_servicio", servicio.nombre_servicio)
        servicio.descripcion_servicio = request.POST.get("descripcion_servicio", servicio.descripcion_servicio)
        servicio.servicio_estado = request.POST.get("servicio_estado") == "True"  # Convertir a booleano

        # Obtener los profesionales seleccionados (IDs enviados desde el formulario)
        profesionales_ids = request.POST.getlist("profesionales[]")
        print("Profesionales seleccionados:", profesionales_ids)

        # Si hay profesionales seleccionados, obtenemos los objetos correspondientes
        if profesionales_ids:
            profesionales = Profesional.objects.filter(id_prof__in=profesionales_ids)
            print("Profesionales encontrados:", profesionales)

            # Actualizar la relación ManyToMany entre el servicio y los profesionales
            servicio.profesionales.set(profesionales)  # Asocia los profesionales seleccionados al servicio
        else:
            servicio.profesionales.clear()  # Si no hay profesionales seleccionados, limpia la relación

        # Guardar los cambios en el servicio
        servicio.save()

        # Obtener los datos actualizados del servicio para enviarlos en la respuesta
        servicio_actualizado = {
            "id_servicio": servicio.id_servicio,
            "nombre_servicio": servicio.nombre_servicio,
            "descripcion_servicio": servicio.descripcion_servicio,
            "servicio_estado": servicio.servicio_estado,
            "profesionales": [
                {"id_prof": prof.id_prof, "nombre_prof": prof.nombre_prof}
                for prof in servicio.profesionales.all()
            ]
        }

        # Devolver los datos actualizados en la respuesta JSON
        return JsonResponse(
            {"status": "success", "message": "Servicio editado correctamente.", "servicio": servicio_actualizado}
        )
    else:
        return JsonResponse({"status": "error", "message": "Método no permitido."})